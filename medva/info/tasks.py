"""
    Celery tasks. Some of them will be launched periodically from admin panel via django-celery-beat
"""

import time

from inmain.celery import app
from celery.utils.log import get_task_logger
from celery import shared_task, current_task
from celery.result import AsyncResult
from time import sleep
import redis
import nltk
from nltk.corpus import stopwords
from nltk.stem import SnowballStemmer
import pymorphy2
import itertools
import redis
import json
from info.models import *
from openpyxl import load_workbook
import io
from datetime import datetime
import random

logger = get_task_logger(__name__)


def getdataset(file_content):
    # Загрузка файла XLSX
    workbook = load_workbook(filename=io.BytesIO(file_content), read_only=True)



    # Выбор активного листа (можно выбрать конкретный лист по имени)
    sheet = workbook.active

    # Создание словаря для хранения данных
    data_dict = {}
    cnt = 0
    cnt2 = 0
    cnt3 = 0

    dataset_list = []

    # Итерация по строкам в листе
    for row in sheet.iter_rows(values_only=True):
        if cnt == 0:
            dataset_dict = {}
            for i in range(len(row)):
                dataset_dict[i] = row[i].strip()
            cnt += 1
        else:
            dataset_list.append(dict(zip([j for j in range(len(row))], ["" for j in range(len(row))])))
            for i in range(len(row)):
                list_of_appointment = str(row[i]).split("\n")
                dataset_list[-1][i] = []
                if len(list_of_appointment) >= 1:
                    for j in range(len(list_of_appointment)):
                        if list_of_appointment[j].strip() != "":
                            dataset_list[-1][i].append(list_of_appointment[j].strip())
    return dataset_list


def lemma_word_pymorphy(word, morph):
    return morph.normal_forms(word)[0]


def get_new_lemma(all_data, morph, stop_words):
    temp_list = all_data.replace("-"," ").replace(","," ").replace("."," ").replace(";"," ").replace("/"," ").replace("–"," ").replace("("," ").replace(")"," ").replace("& "," ").lower()
    tokens = nltk.word_tokenize(temp_list, language='russian')
    tokens = list(set(tokens))
    filtered_tokens = [lemma_word_pymorphy(token,morph) for token in tokens if token not in stop_words]
    return filtered_tokens


def redisgettotal(smth_text, r, morph, stop_words):
    score = {}
    lemmas = get_new_lemma(smth_text, morph, stop_words)
    for elem in lemmas:
        if r.exists(elem):
            serialized_a = r.get(elem)

            # Deserialize the object from JSON
            deserialized_a = json.loads(serialized_a)
            for nmu in deserialized_a[1]:
                if nmu in score:
                    score[nmu] += deserialized_a[0]
                else:
                    score[nmu] = deserialized_a[0]
    return score


def redisgetscoring(smth_text, r, morph, stop_words):
    scoretotal = redisgettotal(smth_text, r, morph, stop_words)
    sorted_dict = dict(sorted(scoretotal.items(), key=lambda x: x[1],  reverse=True))

    top_20 = dict(itertools.islice(sorted_dict.items(), 20))
    return top_20


def createrescription(protocol_id, mkb):
    protocol = Protocol.objects.get(protocol_id=protocol_id)
    if not MedicalStandart.objects.filter(medicalstandartdiagnosis__diagnosis__mkb10=mkb, type=2).exists():
        standarts = MedicalStandart.objects.filter(medicalstandartdiagnosis__diagnosis__mkb10=mkb, type=1).first()
        prescriptionstandarts = PrescriptionMedicalStandart.objects.filter(medical_standart=standarts)
    else:
        standarts = MedicalStandart.objects.get(medicalstandartdiagnosis__diagnosis__mkb10=mkb, type=2)
        prescriptionstandarts = PrescriptionMedicalStandart.objects.filter(medical_standart=standarts)
    standlist=[]
    for standart in prescriptionstandarts:
        newprescription = ProtocolPrescriptionStandart(protocol=protocol, prescription_standart=standart)
        newprescription.save()
        standlist.append(newprescription)
    return standlist


@app.task(ignore_result=True)
def do_work2(data_task):

    r = redis.StrictRedis(
        host="",
        port=6380,
        password="",
        ssl=True,
        ssl_ca_certs='YandexInternalRootCA.crt'
    )

    nltk.download('punkt')
    nltk.download('stopwords')
    stop_words = set(stopwords.words('russian'))
    morph = pymorphy2.MorphAnalyzer()
    """ Get some rest, asynchronously, and update the state all the time """

    clinic = data_task[0]
    filekit = ProtocolFileKit.objects.get(file_kit_id=data_task[1])
    file_in_memory = filekit.file
    file_content = file_in_memory.read()
    data = getdataset(file_content)

    for dataset_list in data:
        print(dataset_list[3][0])
        prescriptlist = dataset_list[7]
        mkb = dataset_list[3][0]
        if Diagnosis.objects.filter(mkb10=mkb).exists():
            diag = Diagnosis.objects.get(mkb10=mkb)
            if MedicalStandart.objects.filter(medicalstandartdiagnosis__diagnosis__mkb10=mkb).exists():
                cnt = 0
                total = prescriptlist
                if not MedicalStandart.objects.filter(medicalstandartdiagnosis__diagnosis__mkb10=mkb,type=2).exists():
                    standarts = MedicalStandart.objects.filter(medicalstandartdiagnosis__diagnosis__mkb10=mkb,type=1).first()
                    prescriptionstandarts = PrescriptionMedicalStandart.objects.filter(medical_standart=standarts)
                else:
                    standarts = MedicalStandart.objects.get(medicalstandartdiagnosis__diagnosis__mkb10=mkb,type=2)
                    prescriptionstandarts = PrescriptionMedicalStandart.objects.filter(medical_standart=standarts)
                specdeparts={
                    'врач-кардиолог': 'Кардиологическое отделение',
                    'врач-невролог': 'Неврологическое отделение',
                    'врач-оториноларинголог': 'Оториноларингологическое отделение',
                }

                doctornames = [
                    'Шестаков Рудольф',
                    'Костин Моисей',
                    'Крюков Геннадий',
                    'Логинов Авраам',
                    'Зайцев Павел',
                    'Кириллов Касьян',
                    'Анисимов Федор'
                ]

                patientnames = [
                    'Шарапов Афанасий',
                    'Громов Всеволод',
                    'Капустин Демьян',
                    'Якушев Флор',
                    'Кузнецов Ян',
                    'Кабанов Афанасий',
                    'Филиппов Степан'
                ]

                if Specialization.objects.filter(name=dataset_list[6][0].lower()).exists():
                    specialization = Specialization.objects.get(name=dataset_list[6][0].lower())
                else:
                    specialization = Specialization(name=dataset_list[6][0].lower())
                    specialization.save()

                clinic_depart = ClinicDepartment.objects.get(clinic=clinic, department__name=specdeparts[dataset_list[6][0]])


                if Doctor.objects.filter(doctordepartment__clinic_department=clinic_depart, doctorspecialization__specialization__name=dataset_list[6][0].lower()):
                    doctorlist = list(Doctor.objects.filter(doctordepartment__clinic_department=clinic_depart, doctorspecialization__specialization__name=dataset_list[6][0].lower()))
                    doctor = random.choice(doctorlist)
                else:
                    randname = random.choice(doctornames)
                    randid = random.randint(100, 1000)
                    if not Doctor.objects.filter(doctor_id=randid).exists():
                        doctor = Doctor(name=randname, id=randid)
                    else:
                        randid = random.randint(100, 1000)
                        doctor = Doctor(name=randname, id=randid)
                    doctor.save()
                    newspec = DoctorSpecialization(specialization=specialization, doctor=doctor)
                    newspec.save()
                    docdep = DoctorDepartment(doctor=doctor, clinic_department=clinic_depart)
                    docdep.save()

                date_obj = datetime.strptime(dataset_list[5][0], '%d.%m.%Y').date()

                if not Patient.objects.filter(id=dataset_list[2][0]):
                    randomnamepatient = random.choice(patientnames)
                    patient = Patient(name=randomnamepatient, id=dataset_list[2][0], sex=dataset_list[0][0])
                    patient.save()
                else:
                    patient = Patient.objects.get(id=dataset_list[2][0])

                newprotocol = Protocol(consultation_date=date_obj,
                                       patient=patient,
                                       doctor=doctor,
                                       clinic_department=clinic_depart,
                                       specialization=specialization,
                                       file_kit=filekit)
                newprotocol.save()



                newdiagprotocol = ProtocolDiagnosis(protocol=newprotocol, diagnosis=diag)
                newdiagprotocol.save()

                standlist = createrescription(protocol_id=newprotocol.protocol_id, mkb=dataset_list[3][0])

                index1=0
                index2=0
                index3=0
                for smth_text in total:
                    prescriptionreal = ProtocolPrescriptionReal(protocol=newprotocol,
                                                                real_text=smth_text)

                    scorenmu1 = redisgetscoring(smth_text, r, morph, stop_words)

                    scorenmu = list(scorenmu1.keys())
                    prescriptneed = []
                    prescriptextra = []
                    for nmustandart in prescriptionstandarts:
                        if not nmustandart.actuality < 1:
                            prescriptneed.append(nmustandart.prescription.nmu_code)
                        else:
                            prescriptextra.append(nmustandart.prescription.nmu_code)
                    prescriptionstandartstotal = prescriptneed+prescriptextra

                    res=''
                    is_find_need = False

                    findlist=[]

                    for nmu in scorenmu:
                        if not is_find_need:
                            if nmu in prescriptneed:
                                prescriptionreal.is_excess=False
                                index3 += 1
                                prescriptionreal.prescription=Prescription.objects.get(nmu_code=nmu)
                                prescriptionreal.show_nmu=True
                                is_find_need = True
                                hasnotnmu=False
                                res='find!!!!!!'
                                findlist.append(nmu)

                                # print(scorenmu1)
                                # print(nmu, scorenmu1[nmu])
                            elif nmu in prescriptextra:
                                res='ok'
                                prescriptionreal.prescription = Prescription.objects.get(nmu_code=nmu)
                                prescriptionreal.show_nmu = True
                                is_find_need = True
                                hasnotnmu = False
                                prescriptionreal.is_excess=False
                                findlist.append(nmu)
                                index3 += 1

                            else:
                                index1 +=1
                                hasnotnmu = True
                                prescriptionreal.is_excess=True
                                prescriptionreal.show_nmu = True

                    if hasnotnmu:
                        prescriptionreal.prescription = Prescription.objects.get(pk=100)
                        prescriptionreal.show_nmu = False
                    prescriptionreal.save()

                    lacked = []


                    for stand in ProtocolPrescriptionStandart.objects.filter(protocol=newprotocol):
                        if stand.prescription_standart.prescription.nmu_code in findlist:
                            stand.is_lacks=False
                            stand.save()
                        else:
                            index2+=1




                    # print(smth_text)
                    print(res)
                    # print('------------')

                newprotocol.index1=index1
                newprotocol.index2=index2
                newprotocol.index3=index3
                newprotocol.save()
                cnt += 1
                i=int(cnt/len(total))
                # logger.info(i)
                # current_task.update_state(state='PROGRESS',
                #     meta={'current': i, 'total': 100})
            else:
                print('not standart')
        else:
            print('not mkb')